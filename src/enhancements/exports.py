"""
Export functionality for the Smart Bus Dashboard.

This module provides comprehensive data export capabilities including
CSV, PDF, Excel formats, and dashboard state sharing.
"""

import streamlit as st
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass
from datetime import datetime
import io
import base64
import json
from pathlib import Path
import zipfile
import tempfile

# Import export libraries
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.config.settings import config_manager


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    include_metadata: bool = True
    include_charts: bool = True
    include_summary: bool = True
    compress_output: bool = False
    timestamp_filename: bool = True
    custom_filename: Optional[str] = None


@dataclass
class ExportResult:
    """Result of an export operation."""
    success: bool
    filename: str
    file_size: int
    format: str
    download_data: Optional[bytes] = None
    error_message: Optional[str] = None


class ExportController:
    """Comprehensive export functionality controller."""
    
    def __init__(self):
        self.export_preferences = config_manager.user_preferences.export_preferences
        self.temp_dir = Path(tempfile.gettempdir()) / "smart_bus_exports"
        self.temp_dir.mkdir(exist_ok=True)
    
    def export_to_csv(self, data: pd.DataFrame, filename: str = None, 
                     config: ExportConfig = None) -> ExportResult:
        """
        Export data to CSV format.
        
        Args:
            data: DataFrame to export
            filename: Custom filename (optional)
            config: Export configuration
            
        Returns:
            ExportResult with download data
        """
        try:
            config = config or ExportConfig()
            
            # Generate filename
            if not filename:
                # Check for custom filename in config
                if config.custom_filename:
                    filename = config.custom_filename
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if config.timestamp_filename else ""
                    filename = f"bus_data_{timestamp}.csv" if timestamp else "bus_data.csv"
            
            # Prepare data
            export_data = data.copy()
            
            # Add metadata if requested
            if config.include_metadata:
                metadata_rows = self._create_metadata_rows()
                # Add metadata as comments (CSV doesn't support metadata directly)
                # We'll add it as a separate sheet in Excel, but for CSV, we'll skip it
                pass
            
            # Convert to CSV
            csv_buffer = io.StringIO()
            export_data.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue().encode('utf-8')
            
            return ExportResult(
                success=True,
                filename=filename,
                file_size=len(csv_data),
                format="CSV",
                download_data=csv_data
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="CSV",
                error_message=str(e)
            )
    
    def export_to_excel(self, data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                       filename: str = None, config: ExportConfig = None) -> ExportResult:
        """
        Export data to Excel format with multiple sheets and formatting.
        
        Args:
            data: DataFrame or dict of DataFrames to export
            filename: Custom filename (optional)
            config: Export configuration
            
        Returns:
            ExportResult with download data
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="Excel",
                error_message="openpyxl library not available"
            )
        
        try:
            config = config or ExportConfig()
            
            # Generate filename
            if not filename:
                # Check for custom filename in config
                if config.custom_filename:
                    filename = config.custom_filename
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if config.timestamp_filename else ""
                    filename = f"bus_dashboard_{timestamp}.xlsx" if timestamp else "bus_dashboard.xlsx"
            
            # Create Excel buffer
            excel_buffer = io.BytesIO()
            
            # Handle single DataFrame or multiple DataFrames
            if isinstance(data, pd.DataFrame):
                data_dict = {"Data": data}
            else:
                data_dict = data
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Write data sheets
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format the sheet
                    worksheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(worksheet, df)
                
                # Add metadata sheet if requested
                if config.include_metadata:
                    metadata_df = self._create_metadata_dataframe()
                    metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
                
                # Add summary sheet if requested
                if config.include_summary and len(data_dict) > 1:
                    summary_df = self._create_summary_dataframe(data_dict)
                    summary_df.to_excel(writer, sheet_name="Summary", index=False)
            
            excel_data = excel_buffer.getvalue()
            
            return ExportResult(
                success=True,
                filename=filename,
                file_size=len(excel_data),
                format="Excel",
                download_data=excel_data
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="Excel",
                error_message=str(e)
            )
    
    def export_to_pdf(self, data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                     title: str = "Smart Bus Dashboard Report",
                     filename: str = None, config: ExportConfig = None) -> ExportResult:
        """
        Export data to PDF format with formatting and charts.
        
        Args:
            data: DataFrame or dict of DataFrames to export
            title: Report title
            filename: Custom filename (optional)
            config: Export configuration
            
        Returns:
            ExportResult with download data
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="PDF",
                error_message="reportlab library not available"
            )
        
        try:
            config = config or ExportConfig()
            
            # Generate filename
            if not filename:
                # Check for custom filename in config
                if config.custom_filename:
                    filename = config.custom_filename
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if config.timestamp_filename else ""
                    filename = f"bus_report_{timestamp}.pdf" if timestamp else "bus_report.pdf"
            
            # Create PDF buffer
            pdf_buffer = io.BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 20))
            
            # Metadata section
            if config.include_metadata:
                story.append(Paragraph("Report Information", styles['Heading2']))
                metadata = self._create_metadata_rows()
                for key, value in metadata.items():
                    story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Handle single DataFrame or multiple DataFrames
            if isinstance(data, pd.DataFrame):
                data_dict = {"Data": data}
            else:
                data_dict = data
            
            # Add data sections
            for section_name, df in data_dict.items():
                # Section header
                story.append(Paragraph(section_name, styles['Heading2']))
                story.append(Spacer(1, 10))
                
                # Summary statistics
                if config.include_summary:
                    summary_stats = self._get_dataframe_summary(df)
                    story.append(Paragraph("Summary Statistics:", styles['Heading3']))
                    for stat, value in summary_stats.items():
                        story.append(Paragraph(f"<b>{stat}:</b> {value}", styles['Normal']))
                    story.append(Spacer(1, 10))
                
                # Data table (first 20 rows)
                story.append(Paragraph("Data Preview:", styles['Heading3']))
                table_data = self._prepare_table_data(df.head(20))
                
                if table_data:
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                
                story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            pdf_data = pdf_buffer.getvalue()
            
            return ExportResult(
                success=True,
                filename=filename,
                file_size=len(pdf_data),
                format="PDF",
                download_data=pdf_data
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="PDF",
                error_message=str(e)
            )
    
    def export_dashboard_state(self, dashboard_state: Dict[str, Any], 
                             filename: str = None) -> ExportResult:
        """
        Export complete dashboard state as JSON.
        
        Args:
            dashboard_state: Current dashboard state
            filename: Custom filename (optional)
            
        Returns:
            ExportResult with download data
        """
        try:
            # Generate filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"dashboard_state_{timestamp}.json"
            
            # Serialize dashboard state
            state_json = json.dumps(dashboard_state, indent=2, default=str)
            state_data = state_json.encode('utf-8')
            
            return ExportResult(
                success=True,
                filename=filename,
                file_size=len(state_data),
                format="JSON",
                download_data=state_data
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                filename="",
                file_size=0,
                format="JSON",
                error_message=str(e)
            )
    
    def generate_shareable_link(self, dashboard_state: Dict[str, Any]) -> str:
        """
        Generate a shareable link for the current dashboard state.
        
        Args:
            dashboard_state: Current dashboard state to encode
            
        Returns:
            Base64 encoded shareable link parameter
        """
        try:
            # Serialize dashboard state
            state_json = json.dumps(dashboard_state, default=str)
            
            # Encode to base64 for URL safety
            state_bytes = state_json.encode('utf-8')
            state_b64 = base64.b64encode(state_bytes).decode('utf-8')
            
            # Create shareable link parameter
            # In a real application, this would be a full URL with the state parameter
            # For now, we return the encoded state that can be appended to a URL
            return state_b64
            
        except Exception as e:
            return f"Error generating link: {str(e)}"
    
    def batch_export(self, datasets: Dict[str, pd.DataFrame], 
                    formats: List[str] = None, 
                    config: ExportConfig = None) -> Dict[str, ExportResult]:
        """
        Export multiple datasets in multiple formats.
        
        Args:
            datasets: Dictionary of dataset names to DataFrames
            formats: List of formats to export (default: ['csv', 'excel', 'pdf'])
            config: Export configuration
            
        Returns:
            Dictionary mapping format names to ExportResults
        """
        if formats is None:
            formats = ['csv', 'excel', 'pdf']
        
        config = config or ExportConfig()
        results = {}
        
        # Export in each requested format
        for format_type in formats:
            format_lower = format_type.lower()
            
            if format_lower == 'csv':
                # For CSV, export each dataset separately
                csv_results = {}
                for name, df in datasets.items():
                    filename = f"{name.lower().replace(' ', '_')}.csv"
                    result = self.export_to_csv(df, filename=filename, config=config)
                    csv_results[name] = result
                
                # If multiple datasets, create a zip file
                if len(datasets) > 1 and config.compress_output:
                    zip_result = self._create_zip_archive(csv_results, "batch_export_csv.zip")
                    results['csv'] = zip_result
                else:
                    # Return the first result for single dataset
                    results['csv'] = list(csv_results.values())[0] if csv_results else ExportResult(
                        success=False, filename="", file_size=0, format="CSV",
                        error_message="No datasets to export"
                    )
            
            elif format_lower == 'excel':
                # Excel can handle multiple sheets natively
                result = self.export_to_excel(datasets, config=config)
                results['excel'] = result
            
            elif format_lower == 'pdf':
                # PDF can handle multiple sections
                result = self.export_to_pdf(datasets, title="Batch Export Report", config=config)
                results['pdf'] = result
        
        return results
    
    def _create_zip_archive(self, export_results: Dict[str, ExportResult], 
                           zip_filename: str) -> ExportResult:
        """
        Create a zip archive from multiple export results.
        
        Args:
            export_results: Dictionary of export results
            zip_filename: Name for the zip file
            
        Returns:
            ExportResult for the zip archive
        """
        try:
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for name, result in export_results.items():
                    if result.success and result.download_data:
                        zip_file.writestr(result.filename, result.download_data)
            
            zip_data = zip_buffer.getvalue()
            
            return ExportResult(
                success=True,
                filename=zip_filename,
                file_size=len(zip_data),
                format="ZIP",
                download_data=zip_data
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                filename=zip_filename,
                file_size=0,
                format="ZIP",
                error_message=str(e)
            )
    
    def create_download_link(self, export_result: ExportResult) -> str:
        """
        Create a download link for exported data.
        
        Args:
            export_result: Result from export operation
            
        Returns:
            Base64 encoded download link
        """
        if not export_result.success or not export_result.download_data:
            return ""
        
        b64_data = base64.b64encode(export_result.download_data).decode()
        
        # Determine MIME type
        mime_types = {
            "CSV": "text/csv",
            "Excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "PDF": "application/pdf",
            "JSON": "application/json"
        }
        
        mime_type = mime_types.get(export_result.format, "application/octet-stream")
        
        return f"data:{mime_type};base64,{b64_data}"
    
    def render_export_controls(self, data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                             section_name: str = "Data") -> None:
        """
        Render export controls in Streamlit interface.
        
        Args:
            data: Data to export
            section_name: Name of the section being exported
        """
        st.markdown("### 📥 Export Options")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📄 Export CSV", use_container_width=True):
                result = self.export_to_csv(data if isinstance(data, pd.DataFrame) else list(data.values())[0])
                if result.success:
                    st.download_button(
                        label="Download CSV",
                        data=result.download_data,
                        file_name=result.filename,
                        mime="text/csv"
                    )
                else:
                    st.error(f"Export failed: {result.error_message}")
        
        with col2:
            if st.button("📊 Export Excel", use_container_width=True):
                result = self.export_to_excel(data)
                if result.success:
                    st.download_button(
                        label="Download Excel",
                        data=result.download_data,
                        file_name=result.filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error(f"Export failed: {result.error_message}")
        
        with col3:
            if st.button("📑 Export PDF", use_container_width=True):
                result = self.export_to_pdf(data, title=f"{section_name} Report")
                if result.success:
                    st.download_button(
                        label="Download PDF",
                        data=result.download_data,
                        file_name=result.filename,
                        mime="application/pdf"
                    )
                else:
                    st.error(f"Export failed: {result.error_message}")
    
    def _format_excel_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """Format Excel worksheet with styling."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metadata_rows(self) -> Dict[str, str]:
        """Create metadata information."""
        return {
            "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Dashboard Version": config_manager.system_config.version,
            "Application": config_manager.system_config.app_name,
            "Export Format": "Multiple formats supported",
            "Data Source": "Smart Bus Scheduling System"
        }
    
    def _create_metadata_dataframe(self) -> pd.DataFrame:
        """Create metadata as DataFrame."""
        metadata = self._create_metadata_rows()
        return pd.DataFrame(list(metadata.items()), columns=["Property", "Value"])
    
    def _create_summary_dataframe(self, data_dict: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Create summary DataFrame for multiple datasets."""
        summary_data = []
        
        for name, df in data_dict.items():
            summary_data.append({
                "Dataset": name,
                "Rows": len(df),
                "Columns": len(df.columns),
                "Memory Usage (KB)": round(df.memory_usage(deep=True).sum() / 1024, 2),
                "Numeric Columns": len(df.select_dtypes(include=[np.number]).columns),
                "Text Columns": len(df.select_dtypes(include=['object']).columns)
            })
        
        return pd.DataFrame(summary_data)
    
    def _get_dataframe_summary(self, df: pd.DataFrame) -> Dict[str, str]:
        """Get summary statistics for DataFrame."""
        return {
            "Total Rows": str(len(df)),
            "Total Columns": str(len(df.columns)),
            "Memory Usage": f"{df.memory_usage(deep=True).sum() / 1024:.2f} KB",
            "Missing Values": str(df.isnull().sum().sum()),
            "Numeric Columns": str(len(df.select_dtypes(include=[np.number]).columns)),
            "Date Columns": str(len(df.select_dtypes(include=['datetime']).columns))
        }
    
    def _prepare_table_data(self, df: pd.DataFrame) -> List[List[str]]:
        """Prepare DataFrame for PDF table."""
        if df.empty:
            return []
        
        # Convert all data to strings and handle long text
        table_data = []
        
        # Headers
        headers = [str(col)[:20] + "..." if len(str(col)) > 20 else str(col) for col in df.columns]
        table_data.append(headers)
        
        # Data rows
        for _, row in df.iterrows():
            row_data = []
            for value in row:
                str_value = str(value)
                # Truncate long values
                if len(str_value) > 20:
                    str_value = str_value[:17] + "..."
                row_data.append(str_value)
            table_data.append(row_data)
        
        return table_data


# Global export controller instance
export_controller = ExportController()